# Mini Project 1
# Student GPA and Ranking
# 22091010303
# Emir Kaan Korkmaz
#
# Some parts of this code are inspired by:
#   - YouTube video (file dialog and Excel loading):
#     "Python Tkinter Tutorial" - https://youtu.be/oZpTv6Z629c?si=UPPp7aAgn5jc3eHu
#   - YouTube video (openpyxl Excel reading):
#     "Python openpyxl tutorial" - https://youtu.be/718edSNvKLA?si=eUfctgZ82eYbnzn5
#
# I completed everything else using the lecture notes and my previous knowledge.



import openpyxl
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from operator import add
from tkinter import messagebox
 


#create function and my source is https://youtu.be/oZpTv6Z629c?si=UPPp7aAgn5jc3eHu
def load_data():
    
    global names, name, student_id, student_ids, row, surname, surnames, answer_list, rank_list
    names = [] #we need this for name list
    student_ids = [] #we need this for student ID list
    surnames = [] #we need this for surname list
    physics_list=[]
    calculus_list=[]
    advanced_programming_list=[]
    chemistry_list=[]
    types = [ ("Text Files", "*.xlsx"), ("All Files", "*.*")]            
    file_path = filedialog.askopenfilename(filetypes=types, initialdir=".", title="Load Data")
    messagebox.showinfo("Success", "The file has been successfully loaded!")

    #we bring data from "excel" and my source is https://youtu.be/718edSNvKLA?si=eUfctgZ82eYbnzn5
    book = load_workbook(file_path)
    sheet = book.active
    row = sheet.max_row

  
    for i in range(2, row + 1):
        name = sheet.cell(row=i, column=1).value # we have names
        names.append(name)
        surname = sheet.cell(row=i, column=2).value # we have surnames
        surnames.append(surname)
        student_id = sheet.cell(row=i, column=3).value # we have student ID
        student_ids.append(student_id)

    for i in range(2, row + 1):
        physics = sheet.cell(row=i, column=4).value
        physics_list.append(physics)
        calculus = sheet.cell(row=i, column=5).value
        calculus_list.append(calculus)       
        advanced_programming = sheet.cell(row=i, column=6).value
        advanced_programming_list.append(advanced_programming)
        chemistry = sheet.cell(row=i, column=7).value
        chemistry_list.append(chemistry)

        answer_list = []

    # calculate GPA
    try:
        for i in range(len(physics_list)):
            physics_score = physics_list[i] * 0.25
            calculus_score = calculus_list[i] * 0.25
            advanced_programming_score = advanced_programming_list[i] * 0.30
            chemistry_score = chemistry_list[i] * 0.20
            total = physics_score + calculus_score + advanced_programming_score + chemistry_score
            total = round(total, 2)
            answer_list.append(total) 

    except:
        print("Error: Failed to calculate GPA due to invalid values.")
        return   

    sorted_gpas = sorted(answer_list, reverse=True)
    rank_list = []

    for gpa in answer_list:
        rank = sorted_gpas.index(gpa) + 1
        rank_list.append(rank)


def display():

    global names, name, student_id, student_ids, row, surname, surnames, answer_list, rank_list

    try:   
        if len(student_ids) == 0:
            raise Exception("File not loaded")
        
        for i in range(len(student_ids)):
            if entry2.get() == str(student_ids[i]):
                show_name_surname = f"{names[i]} {surnames[i]}"
                label33.config(text=show_name_surname)
                show_gpa = f"{answer_list[i]}"
                label44.config(text=show_gpa)
                rank_value = f"{rank_list[i]}"
                label55.config(text=rank_value)

    except Exception:
        print("Warning: You must load a file first!")

def export():
    # use global widgets
    global combo_box, label33, label44, label55, name, surname

    selected_type = combo_box.get()
    name_surname = label33.cget("text")
    gpa = label44.cget("text")
    rank = label55.cget("text")

    # if nothing displayed yet
    if name_surname == "":
        messagebox.showwarning("Warning", "Please display a student before exporting.")
        return
    
    student_id_text = entry2.get().strip()
    default_filename = f"{student_id_text} {name} {surname}"

    # TEXT EXPORT
    if selected_type == ".txt":
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt")],
            initialdir=".",
            initialfile=default_filename
        )
        
        if file_path:  # user didn't cancel
            with open(file_path, "w") as f:
                f.write(f"Name Surname: {name_surname}\n")
                f.write(f"GPA: {gpa}\n")
                f.write(f"Rank: {rank}\n")
            messagebox.showinfo("Success", "The file has been successfully exported!")

    # EXCEL EXPORT
    elif selected_type == ".xls":
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialdir=".",
            initialfile=default_filename
        )
        if file_path:
            wb = Workbook()
            ws = wb.active

            # headers
            ws["A1"] = "Name Surname"
            ws["B1"] = "GPA"
            ws["C1"] = "Rank"

            # data
            ws["A2"] = name_surname
            ws["B2"] = float(gpa) 
            ws["C2"] = rank

            wb.save(file_path)
            messagebox.showinfo("Success", "The file has been successfully exported!")

    else:
        messagebox.showwarning("Warning", "Please select a valid file type (.txt or .xls).")

def clear():  
 
    entry2.delete(0, END) #clear id 

    label33.config(text="") #clear labels
    label44.config(text="")
    label55.config(text="")


    combo_box.set(".txt") #reset combo box

    messagebox.showinfo("Cleared", "All fields have been cleared.")



# Define the main window
master = Tk()
master.title("Mini Project 01")
master.geometry("350x300")


# Create widgets
title_label = Label(master, text="Student GPA and Ranking")
label1 = Label(master, text="Open File")
button1 = Button(master, text="Browse", command =load_data)
label2 = Label(master, text="ID:")
entry2 = Entry(master)
label3 = Label(master, text="Name Surname:")
label33 = Label(master, text="")
label4 = Label(master, text="GPA:")
label44 = Label(master, text="")
label5 = Label(master, text="Rank:")
label55 = Label(master, text="")
label6 = Label(master, text="Please select file type: ")
combo_box = ttk.Combobox(master, values=[".txt", ".xls"])
button2 = Button(master, text="Display", command=display)
button3 = Button(master, text="Export", command=export)
button4 = Button(master, text="Clear", command=clear)


# Place widgets using the grid layout
title_label.grid(row=0, column=0, columnspan=3, pady=10)
label1.grid(row=1, column=0, sticky=W)
button1.grid(row=1, column=1, columnspan=2, sticky=W+E)
label2.grid(row=2, column=0, columnspan=2, sticky=W) 
entry2.grid(row=2, column=1, columnspan=2, sticky=W+E) 
label3.grid(row=3, column=0, sticky=W)
label33.grid(row=3, column=1)
label4.grid(row=4, column=0, sticky=W)
label44.grid(row=4, column=1)
label5.grid(row=5, column=0, sticky=W)
label55.grid(row=5, column=1)
label6.grid(row=6, column=0, sticky=W)
combo_box.grid(row=6, column=1, columnspan=2)
combo_box.set(".txt")
button2.grid(row=7, column=0, sticky=W+E+S+N)
button3.grid(row=7, column=1,  sticky=W+E)
button4.grid(row=7, column=2, sticky=E+W)


# Run the Tkinter main loop
master.mainloop()

